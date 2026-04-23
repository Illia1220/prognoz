from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import io
import json
import tempfile

import numpy as np
import pandas as pd
from db import supabase
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

app = Flask(__name__)
CORS(app)


def calculate_metrics(df):
    df = df.fillna(0)

    df["ctr"] = df["clicks"] / df["impressions"].replace(0, np.nan)
    df["cpm"] = df["spend"] / df["impressions"].replace(0, np.nan) * 1000
    df["cpa"] = df["spend"] / df["conversions"].replace(0, np.nan)
    df["roi"] = df["revenue"] / df["spend"].replace(0, np.nan)

    df = df.replace([np.inf, -np.inf], 0)
    df = df.fillna(0)
    return df


def load_ads_dataframe():
    response = supabase.table("ads_data").select("*").execute()
    data = response.data or []

    if not data:
      return pd.DataFrame(), []

    df = pd.DataFrame(data)
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df = df.dropna(subset=["date"])

    if df.empty:
        return pd.DataFrame(), data

    df = calculate_metrics(df)
    return df, data


def safe_percent_delta(current, previous):
    if previous in (0, None) or pd.isna(previous):
        return 0.0
    return float(((current - previous) / abs(previous)) * 100)


def build_monthly_metrics(df):
    monthly = df.groupby(df["date"].dt.to_period("M")).agg({
        "spend": "sum",
        "revenue": "sum",
        "clicks": "sum",
        "impressions": "sum",
        "conversions": "sum",
    }).reset_index()

    monthly["date"] = monthly["date"].astype(str)
    monthly["roi"] = (monthly["revenue"] / monthly["spend"].replace(0, np.nan)).fillna(0)
    monthly["ctr"] = (monthly["clicks"] / monthly["impressions"].replace(0, np.nan)).fillna(0)
    monthly["cpa"] = (monthly["spend"] / monthly["conversions"].replace(0, np.nan)).fillna(0)
    monthly = monthly.replace([np.inf, -np.inf], 0).fillna(0)
    return monthly


def build_campaign_metrics(df):
    campaigns = df.groupby("campaign").agg({
        "spend": "sum",
        "revenue": "sum",
        "clicks": "sum",
        "impressions": "sum",
        "conversions": "sum",
    }).reset_index()

    campaigns["roi"] = (campaigns["revenue"] / campaigns["spend"].replace(0, np.nan)).fillna(0)
    campaigns["ctr"] = (campaigns["clicks"] / campaigns["impressions"].replace(0, np.nan)).fillna(0)
    campaigns["cpa"] = (campaigns["spend"] / campaigns["conversions"].replace(0, np.nan)).fillna(0)
    campaigns = campaigns.replace([np.inf, -np.inf], 0).fillna(0)
    return campaigns


def build_forecast_payload(df):
    if df.empty:
        return {
            "next_month_roi": 0,
            "roi_trend": 0,
            "recommended_spend": 0,
            "monthly": [],
            "forecast_point": None,
            "comparison": {
                "current_period": None,
                "previous_period": None,
                "roi_delta": 0,
                "spend_delta": 0,
                "cpa_delta": 0,
            },
            "highlights": {
                "top_campaigns": [],
                "risk_campaigns": [],
                "alerts": [],
                "recommendation": {
                    "label": "Hold",
                    "text": "Upload campaign data to receive recommendations.",
                },
            },
        }

    monthly = build_monthly_metrics(df)

    if len(monthly) > 1:
        roi_trend = (monthly["roi"].iloc[-1] - monthly["roi"].iloc[0]) / len(monthly)
    else:
        roi_trend = 0

    avg_roi = monthly["roi"].mean()
    next_month_roi = avg_roi + roi_trend
    last_spend = monthly["spend"].iloc[-1]

    roi_factor = 1.15 if next_month_roi > avg_roi else 0.90 if next_month_roi < avg_roi else 1
    last_ctr = monthly["ctr"].iloc[-1]
    avg_ctr = monthly["ctr"].mean()
    ctr_factor = 1.05 if last_ctr > avg_ctr else 0.95

    recommended_spend = last_spend * roi_factor * ctr_factor
    recommended_spend = min(recommended_spend, last_spend * 1.20)
    recommended_spend = max(recommended_spend, last_spend * 0.70)

    last_month = pd.Period(monthly["date"].iloc[-1], freq="M")
    forecast_point = {
        "date": str(last_month + 1),
        "roi": float(next_month_roi),
    }

    current_period = monthly.iloc[-1].to_dict()
    previous_period = monthly.iloc[-2].to_dict() if len(monthly) > 1 else None

    comparison = {
        "current_period": current_period["date"],
        "previous_period": previous_period["date"] if previous_period else None,
        "roi_delta": safe_percent_delta(current_period["roi"], previous_period["roi"]) if previous_period else 0,
        "spend_delta": safe_percent_delta(current_period["spend"], previous_period["spend"]) if previous_period else 0,
        "cpa_delta": safe_percent_delta(current_period["cpa"], previous_period["cpa"]) if previous_period else 0,
    }

    campaigns = build_campaign_metrics(df)
    top_campaigns = campaigns.sort_values("roi", ascending=False).head(3)
    risk_campaigns = campaigns.sort_values("roi", ascending=True).head(3)

    alerts = []
    if current_period["roi"] < 1:
        alerts.append({
            "title": "ROI pressure",
            "text": "Current month ROI is below break-even. Review weak campaigns first.",
            "tone": "negative",
        })
    if comparison["cpa_delta"] > 15:
        alerts.append({
            "title": "CPA spike",
            "text": f"CPA increased by {comparison['cpa_delta']:.1f}% versus the previous month.",
            "tone": "warn",
        })
    if comparison["spend_delta"] > 25:
        alerts.append({
            "title": "Spend surge",
            "text": f"Spend increased by {comparison['spend_delta']:.1f}%. Double-check scaling quality.",
            "tone": "warn",
        })
    if roi_trend > 0 and next_month_roi > 1:
        alerts.append({
            "title": "Scaling signal",
            "text": "Forecast trend is positive and ROI stays above break-even.",
            "tone": "positive",
        })

    if roi_trend > 0.03 and next_month_roi > 1:
        recommendation = {
            "label": "Scale",
            "text": "Positive momentum and profitable forecast suggest careful spend expansion.",
        }
    elif roi_trend < 0 or next_month_roi < 1:
        recommendation = {
            "label": "Reduce",
            "text": "Forecast weakens. Cut inefficient campaigns and protect margin before scaling again.",
        }
    else:
        recommendation = {
            "label": "Hold",
            "text": "Signals are mixed. Keep budgets stable until the next clear move appears.",
        }

    return {
        "monthly": monthly.to_dict(orient="records"),
        "next_month_roi": float(next_month_roi),
        "roi_trend": float(roi_trend),
        "recommended_spend": float(recommended_spend),
        "forecast_point": forecast_point,
        "comparison": comparison,
        "highlights": {
            "top_campaigns": top_campaigns[["campaign", "roi", "spend", "clicks"]].to_dict(orient="records"),
            "risk_campaigns": risk_campaigns[["campaign", "roi", "spend", "clicks"]].to_dict(orient="records"),
            "alerts": alerts,
            "recommendation": recommendation,
        },
    }


@app.route("/clear", methods=["POST"])
def clear():
    try:
        supabase.table("ads_data").delete().gte("date", "1900-01-01").execute()
        return jsonify({"status": "cleared"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/upload", methods=["POST"])
def upload_csv():
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file provided"}), 400

        file = request.files["file"]
        if file.filename == "":
            return jsonify({"error": "Empty file"}), 400

        content = file.read().decode("utf-8")
        df = pd.read_csv(io.StringIO(content))

        allowed_cols = [
            "date", "campaign", "geo",
            "impressions", "clicks", "spend",
            "conversions", "revenue"
        ]

        df = df[allowed_cols]
        df = df.dropna(how="all")

        for col in ["impressions", "clicks", "spend", "conversions", "revenue"]:
            df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0)

        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        df = df.dropna(subset=["date"])
        df["date"] = df["date"].dt.strftime("%Y-%m-%d")

        df = df.drop_duplicates()
        df = calculate_metrics(df)
        df = df.replace({np.nan: None})

        data = json.loads(df.to_json(orient="records"))
        supabase.table("ads_data").delete().gte("date", "1900-01-01").execute()
        supabase.table("ads_data").insert(data).execute()

        return jsonify({"status": "success", "rows_inserted": len(data)})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/forecast", methods=["GET"])
def forecast():
    try:
        df, _ = load_ads_dataframe()
        return jsonify(build_forecast_payload(df))
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route("/export", methods=["GET"])
def export_excel():
    try:
        df, data = load_ads_dataframe()
        if not data or df.empty:
            return jsonify({"error": "No data"}), 400

        forecast_payload = build_forecast_payload(df)
        campaign_stats = build_campaign_metrics(df)
        global_avg_roi = campaign_stats["roi"].mean()

        campaign_stats["predicted_roi"] = campaign_stats["roi"]
        campaign_stats["recommended_spend"] = campaign_stats["spend"] * (
            campaign_stats["predicted_roi"] / (global_avg_roi if global_avg_roi != 0 else 1)
        )

        wb = Workbook()
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)

        def style_sheet(ws):
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        value = str(cell.value)
                        if len(value) > max_length:
                            max_length = len(value)
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 3

        ws1 = wb.active
        ws1.title = "Raw Data"

        export_df = df.merge(
            campaign_stats[["campaign", "predicted_roi", "recommended_spend"]],
            on="campaign",
            how="left",
        )
        export_df["date"] = pd.to_datetime(export_df["date"]).dt.strftime("%d.%m.%Y %H:%M:%S")

        ws1.append(list(export_df.columns))
        for row in export_df.itertuples(index=False):
            ws1.append(list(row))
        style_sheet(ws1)

        ws2 = wb.create_sheet("Campaign Forecast")
        ws2.append(["Campaign", "ROI", "Predicted ROI", "Recommended Spend"])
        for row in campaign_stats.itertuples(index=False):
            ws2.append([row.campaign, float(row.roi), float(row.predicted_roi), float(row.recommended_spend)])
        style_sheet(ws2)

        ws3 = wb.create_sheet("Summary")
        ws3.append(["Metric", "Value"])
        ws3.append(["Next Month ROI", float(forecast_payload["next_month_roi"])])
        ws3.append(["ROI Trend", float(forecast_payload["roi_trend"])])
        ws3.append(["Recommended Spend", float(forecast_payload["recommended_spend"])])
        ws3.append(["ROI Delta %", float(forecast_payload["comparison"]["roi_delta"])])
        ws3.append(["Spend Delta %", float(forecast_payload["comparison"]["spend_delta"])])
        ws3.append(["CPA Delta %", float(forecast_payload["comparison"]["cpa_delta"])])
        ws3.append(["Recommendation", forecast_payload["highlights"]["recommendation"]["label"]])
        style_sheet(ws3)

        tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        wb.save(tmp.name)
        return send_file(tmp.name, as_attachment=True, download_name="analytics_report.xlsx")
    except Exception as e:
        return jsonify({"error": str(e)}), 500


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=10000)
